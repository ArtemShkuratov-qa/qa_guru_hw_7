import os
import csv
import zipfile
from pypdf import PdfReader
from openpyxl import load_workbook
from conftest import RES_DIR


def test_csv():
    with zipfile.ZipFile(os.path.join(RES_DIR, 'companies.zip')) as zf:
        with zf.open('tmp/import_company_csv.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            second_row = csvreader[1]

            assert second_row[0] == 'OU001'
            assert second_row[2] == 'ЗАО Сервис'
            print('Тест пройден')


def test_pdf():
    with zipfile.ZipFile(os.path.join(RES_DIR, 'companies.zip')) as zf:
        with zf.open('tmp/import_company_pdf.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            number_of_pages = len(reader.pages)
            page = reader.pages[0]
            text = page.extract_text()
            row = list(text.splitlines())

            assert  number_of_pages == 2
            assert row[0] == 'Внешний идентификатор для импорта Вышестоящий отдел'
            assert row[4] == 'OU004'
            print('Тест пройден')


def test_xlsx():
    with zipfile.ZipFile(os.path.join(RES_DIR, 'companies.zip')) as zf:
        with zf.open('tmp/import_company_xlsx.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active

            assert sheet.cell(row=6, column=3). value == 'Филиал 1'
            assert sheet.cell(row=6, column=4).value == 'Челябинск'
            print('Тест пройден')


